from __future__ import annotations
import io
import json
import os
from datetime import datetime
from openpyxl import load_workbook


def _to_str(v) -> str:
    if v is None:
        return ""
    return str(v)


def read_excel_preview(
    file_bytes: bytes,
    sheet_index: int = 0,
    max_rows: int = 20,
) -> tuple[list[str], list[list]]:
    """
    Returns (sheet_names, grid).
    sheet_names: all sheet names in the workbook.
    grid: first max_rows rows as list of lists (all columns, values as strings).
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    idx = max(0, min(sheet_index, len(wb.worksheets) - 1))
    ws = wb.worksheets[idx]
    grid: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        grid.append([_to_str(v) for v in row])
    wb.close()
    return sheet_names, grid


def read_excel_rows(
    file_bytes: bytes,
    column_map: dict,
    required_fields: list[str],
) -> list[dict]:
    """
    Reads the sheet from start_row onward.
    column_map keys: "sheet_index", "start_row", plus one key per field → 1-based column index.
    Returns list of dicts keyed by field name; each dict also has "_row" (1-based row number).
    Skips rows where every required field cell is empty/None.
    """
    sheet_index = int(column_map.get("sheet_index", 0))
    start_row = int(column_map.get("start_row", 2))
    field_cols: dict[str, int] = {
        k: int(v)
        for k, v in column_map.items()
        if k not in ("sheet_index", "start_row") and isinstance(v, (int, float)) and int(v) > 0
    }

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    idx = max(0, min(sheet_index, len(wb.worksheets) - 1))
    ws = wb.worksheets[idx]

    rows: list[dict] = []
    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_num < start_row:
            continue
        row_dict: dict = {"_row": row_num}
        for field_key, col_idx in field_cols.items():
            val = row[col_idx - 1] if col_idx <= len(row) else None
            # Treat Excel error strings as empty
            if isinstance(val, str) and val.startswith("#"):
                val = None
            row_dict[field_key] = val

        # Skip if all required fields are blank
        if required_fields and all(
            not row_dict.get(f) for f in required_fields
        ):
            continue
        rows.append(row_dict)

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Pre-import backup
# ---------------------------------------------------------------------------

def backup_import_tables(
    conn,
    specs: list[dict],
    client_id: int,
    env: str,
) -> list[tuple[str, int | None, str | None]]:
    """
    Snapshot affected tables to JSON files before an import run.
    Files are written in the same format as migration/backup.py so they appear
    in Settings → Backup Manager and can be restored for tables that have a
    direct ClientId column.

    Each spec dict:
      table            – table name
      query            – SELECT to fetch the rows to snapshot
      params           – query params tuple
      client_id_column – (optional) the ClientId column name; used by the
                         restore function in Settings.  Omit or set to None
                         for tables without a direct ClientId column —
                         their backups are saved but cannot be auto-restored.

    Returns list of (file_path, row_count, error_message) tuples.
    error_message is None on success.
    """
    from config import BACKUP_DIR

    env_dir = os.path.join(BACKUP_DIR, env or "unknown")
    os.makedirs(env_dir, exist_ok=True)

    results: list[tuple[str, int | None, str | None]] = []

    for spec in specs:
        table = spec["table"]
        query = spec["query"]
        params = spec["params"]
        cid_col = spec.get("client_id_column", "ClientId")

        try:
            with conn.cursor() as cur:
                cur.execute(query, params)
                rows = cur.fetchall()

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"clt_bkp_{client_id}_{table[:40]}_{ts}.json"
            file_path = os.path.join(env_dir, filename)

            payload = {
                "metadata": {
                    "client_id":        client_id,
                    "original_table":   table,
                    "client_id_column": cid_col or "ClientId",
                    "env":              env,
                    "created_at":       datetime.now().isoformat(),
                    "source":           "data_import",
                    "restorable":       cid_col is not None,
                },
                "rows": rows,
            }

            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, default=str, indent=2)

            results.append((file_path, len(rows), None))

        except Exception as exc:
            results.append(("", None, f"{table}: {exc}"))

    return results
